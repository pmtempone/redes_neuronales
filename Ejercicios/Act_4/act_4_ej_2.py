# levanto el archivo
from openpyxl import load_workbook
import func.som as som
import numpy as np

wb = load_workbook('Ejercicios/Act_4/Drug4.xlsx', read_only=True)
ws = wb.get_sheet_by_name('Sheet1')
ws.max_row
ws.max_column
numerical_data = ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column)
muestra = np.array([[cell.value for cell in row] for row in numerical_data])
# separo los datos de la calse
X = muestra[:, 0:6]
T = muestra[:, 6]
# escalamos los datos
X_escalado = (X - X.min(axis=0)) / (X.max(axis=0) - X.min(axis=0))
# Genero matriz de valores esperados
T_matriz = np.concatenate(([T == 0], [T == 1], [T == 2], [T == 3]), axis=0).astype(int)

# Defino Parámetros de la red
P = X_escalado.T
columnas = 4
filas = 4
alfaInicial = 0.2
vecindad = 3
func_vecindad = 1  ####otro nro es sombrero
sigma = 2  ##ancho del sombrero
ite_reduce = 50
dibujar = True

matriz = np.zeros((200, 200))
for func_vecindad in [0, 1]:
    for vecindad in [2, 3]:
        for sigma in [1, 2, 3]:
            for columnas in [2, 3, 4]:
                (w_O) = som.train(P, filas, columnas, alfaInicial, vecindad, func_vecindad, sigma, ite_reduce, dibujar)

                # una columna por cada dimension de mi dataset
                ganador = []
                for p in range(200):
                    distancias = -np.sqrt(
                        np.sum((w_O - (P[:, p][np.newaxis]) * np.ones((filas * columnas, 1))) ** 2, 1))
                    ganadora = np.argmax(distancias)
                    ganador.append(ganadora)

                for i in range(200):
                    for j in range(200):
                        ##if i!=j: opcional.....la funcion dendograma se encarga de esto
                        if ganador[i] == ganador[j]:
                            matriz[i][j] += 1

matriz2 = matriz



som.dendograma(matriz, T)